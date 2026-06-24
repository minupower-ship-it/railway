import asyncio
import re
import discord
from discord import app_commands
from discord.ext import tasks
from datetime import datetime, time
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import os
import aiohttp
import json
import io
import openpyxl
import asyncpg

load_dotenv()

TOKEN             = os.getenv('DISCORD_TOKEN')
RENDER_URL        = "https://xh-7vlt.onrender.com"
API_SECRET_KEY    = os.getenv('API_SECRET_KEY')
DATABASE_URL       = os.getenv('DATABASE_URL')
LINK_STORE_ID      = 1488022953525248141  # XHouse Dummy 채널 ID (마이그레이션용)
PREVIEW_CHANNEL_ID = 1487501681129422980  # preview 채널
EDMONTON_TZ        = ZoneInfo("America/Edmonton")
SUPPORT_CHANNEL_ID  = int(os.getenv('SUPPORT_CHANNEL_ID', '0'))
TX_CHANNEL_ID       = int(os.getenv('XHOUSE_TX_CHANNEL_ID', '1486794773263024309'))

db_pool      = None
invite_cache = {}  # {guild_id: {invite_code: uses}}

XHOUSE_GUILD_ID_INT  = int(os.getenv('XHOUSE_GUILD_ID', '0'))
XHOUSE_ROLE_ID_INT   = int(os.getenv('XHOUSE_ROLE_ID',  '0'))
VIP_ROLE_ID_INT      = int(os.getenv('XHOUSE_VIP_ROLE_ID', '1519415094235365386'))
VIP_WINDOW_SIZE      = 20  # 채널당 최신 N개 = VIP 전용

intents = discord.Intents.default()
intents.message_content = True
intents.members = True

client = discord.Client(intents=intents)
tree   = app_commands.CommandTree(client)

# ================== 채널 ID ==================
CHANNELS = {
    "🇭🇮🇸🇵🇦🇳🇮🇨":       1487319298681864342,
    "🇼🇭🇮🇹🇪":            1487319326204625047,
    "🇧🇱🇦🇨🇰":            1487319363265626173,
    "🇦🇸🇮🇦🇳-🇴🇹🇭🇪🇷🇸": 1487319260228358174,
    "🇨🇴🇱🇱🇦🇧s":         1489310534544265376,
}

# ================== TXT 채널 매핑 ==================
CHANNEL_MAP = {
    "asian":    1487319260228358174,  # 🇦🇸🇮🇦🇳-🇴🇹🇭🇪🇷🇸
    "hispanic": 1487319298681864342,  # 🇭🇮🇸🇵🇦🇳🇮🇨
    "white":    1487319326204625047,  # 🇼🇭🇮🇹🇪
    "black":    1487319363265626173,  # 🇧🇱🇦🇨🇰
    "collabs":  1489310534544265376,  # 🇨🇴🇱🇱🇦🇧s
}

# ================== DB 초기화 ==================
async def init_db():
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL)
    async with db_pool.acquire() as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS links (
                thread_id BIGINT PRIMARY KEY,
                mega_link TEXT NOT NULL
            )
        ''')
        await conn.execute('ALTER TABLE links ADD COLUMN IF NOT EXISTS vip BOOLEAN DEFAULT FALSE')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS posted_names (
                name TEXT PRIMARY KEY
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS preview_msg (
                id SERIAL PRIMARY KEY,
                message_id BIGINT
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS paid_invites (
                invite_code TEXT PRIMARY KEY,
                session_id  TEXT NOT NULL,
                ref         TEXT,
                used        BOOLEAN DEFAULT FALSE,
                created_at  TIMESTAMP DEFAULT NOW()
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS referral_conversions (
                id           SERIAL PRIMARY KEY,
                session_id   TEXT,
                ref          TEXT,
                discord_id   BIGINT,
                plan         TEXT,
                converted_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        await conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS uq_rc_session ON referral_conversions(session_id)')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS promoter_invites (
                promoter    TEXT PRIMARY KEY,
                invite_code TEXT NOT NULL,
                invite_url  TEXT NOT NULL
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS member_sources (
                discord_id  BIGINT PRIMARY KEY,
                promoter    TEXT NOT NULL,
                joined_at   TIMESTAMP DEFAULT NOW()
            )
        ''')
    print("✅ DB 초기화 완료")

# ================== Dummy 채널 → DB 마이그레이션 ==================
async def migrate_from_dummy():
    async with db_pool.acquire() as conn:
        count = await conn.fetchval('SELECT COUNT(*) FROM links')
        if count > 0:
            print(f"[Migration] 이미 {count}개 링크 있음, 스킵")
            return

    print("[Migration] Dummy 채널에서 링크 가져오는 중...")
    channel = client.get_channel(LINK_STORE_ID)
    if not channel:
        print("[Migration] Dummy 채널 없음")
        return

    migrated = 0
    async for message in channel.history(limit=2000):
        content = message.content or ""
        if " | " in content and not content.startswith("POSTED |") and not content.startswith("PREVIEW_MSG |"):
            parts = content.split(" | ", 1)
            if len(parts) == 2:
                try:
                    thread_id = int(parts[0].strip())
                    mega_link = parts[1].strip()
                    async with db_pool.acquire() as conn:
                        await conn.execute(
                            'INSERT INTO links (thread_id, mega_link) VALUES ($1, $2) ON CONFLICT DO NOTHING',
                            thread_id, mega_link
                        )
                    migrated += 1
                except ValueError:
                    pass
    print(f"[Migration] 완료: {migrated}개 링크 마이그레이션")

# ================== 링크 저장/조회 헬퍼 ==================
async def save_link(thread_id: int, link: str):
    async with db_pool.acquire() as conn:
        await conn.execute(
            'INSERT INTO links (thread_id, mega_link) VALUES ($1, $2) ON CONFLICT (thread_id) DO UPDATE SET mega_link = $2',
            thread_id, link
        )

async def get_link(thread_id: int) -> str:
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT mega_link FROM links WHERE thread_id = $1', thread_id)
    return row['mega_link'] if row else None

# ================== Content Request ==================
class ContentRequestModal(discord.ui.Modal, title="Content Request Form"):
    name    = discord.ui.TextInput(label="Name", placeholder="Creator name", required=True, max_length=100)
    link    = discord.ui.TextInput(label="Link (OF / X / etc.)", placeholder="https://...", required=True)
    comment = discord.ui.TextInput(label="Comment", placeholder="Additional comments", required=False, style=discord.TextStyle.paragraph)

    async def on_submit(self, interaction: discord.Interaction):
        try:
            request_channel = await client.fetch_channel(1486473063649247352)
        except Exception:
            request_channel = None
        if request_channel:
            embed = discord.Embed(title="🆕 New Content Request", description="A new content request has been submitted!", color=0x9b59b6, timestamp=datetime.utcnow())
            embed.add_field(name="Requested By", value=f"{interaction.user.display_name}\n({interaction.user.name})", inline=True)
            embed.add_field(name="Name", value=self.name.value, inline=True)
            embed.add_field(name="Content Link", value=self.link.value, inline=False)
            if self.comment.value and self.comment.value.strip():
                embed.add_field(name="Comment", value=self.comment.value, inline=False)
            embed.set_footer(text=f"Requested at • {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC")
            await request_channel.send(embed=embed)
            await interaction.response.send_message("✅ Your request has been submitted successfully!", ephemeral=True)
        else:
            await interaction.response.send_message("❌ Request channel not found. Please contact an admin.", ephemeral=True)


class RequestButtonView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)
        self.add_item(discord.ui.Button(label="Funding", style=discord.ButtonStyle.link, emoji="💰", url="https://buy.stripe.com/3cI6oJaZo8PS1kF9281ck0m"))

    @discord.ui.button(label="Submit Content Request", style=discord.ButtonStyle.primary, emoji="📩", custom_id="submit_request")
    async def submit(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(ContentRequestModal())


@tree.command(name="setup-request", description="request-form 채널에 버튼 설정")
async def setup_request(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
        return
    embed = discord.Embed(title="Welcome to #request-form!", description="Submit your content requests here.", color=0x2b2d31)
    embed.add_field(name="Request Any Model You Want", value="Want access to exclusive content?\n\nGet premium access at **xhouse.vip** and unlock:\n• Unlimited model requests\n• Exclusive content library\n• Priority updates", inline=False)
    view = RequestButtonView()
    await interaction.response.send_message(embed=embed, view=view)


# ================== Payment System ==================
class PaymentView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="Get Lifetime — $45", style=discord.ButtonStyle.success, emoji="💳", custom_id="xhouse_lifetime")
    async def get_lifetime(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._handle_payment(interaction, 'lifetime')

    @discord.ui.button(label="Get VIP — $80", style=discord.ButtonStyle.primary, emoji="✨", custom_id="xhouse_vip")
    async def get_vip(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._handle_payment(interaction, 'vip')

    async def _handle_payment(self, interaction: discord.Interaction, plan: str):
        await interaction.response.defer(ephemeral=True)
        discord_id = str(interaction.user.id)
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    f"{RENDER_URL}/create-checkout",
                    json={"plan": plan, "discord_id": discord_id},
                    headers={"Content-Type": "application/json", "X-API-Key": API_SECRET_KEY}
                ) as res:
                    data = await res.json()
            if data.get("url"):
                await interaction.followup.send(
                    f"✅ Complete your payment using the link below!\n{data['url']}\n\nYour role will be granted automatically after payment.",
                    ephemeral=True
                )
            else:
                await interaction.followup.send(f"❌ Failed to generate payment link.\nServer response: `{data}`", ephemeral=True)
        except Exception as e:
            print(f"[XHouse Bot] Payment error: {e}")
            await interaction.followup.send("❌ Server error. Please try again later.", ephemeral=True)


@tree.command(name="setup-payment", description="결제 버튼 설정")
async def setup_payment(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
        return
    embed = discord.Embed(
        title="💎 X-House Membership",
        description="Get lifetime access to exclusive content, model packages, and private channels.",
        color=0x7b6eff
    )
    embed.add_field(name="Lifetime — $45", value="• Private request\n• 1,000+ model packages\n• Lifetime access to all channels\n• Priority support", inline=True)
    embed.add_field(name="VIP Lifetime — $80", value="• Everything in Lifetime\n• Unlimited requests for life\n• Early access to new drops\n• Private personal request", inline=True)
    embed.set_footer(text="One-time payment. No subscriptions. No renewals.")
    view = PaymentView()
    await interaction.response.send_message(embed=embed, view=view)


# ================== Post System ==================
class PostModal(discord.ui.Modal, title="New Content Post"):
    post_name = discord.ui.TextInput(label="Name", placeholder="e.g. Poppi Louiz", required=True, max_length=100)
    file_size = discord.ui.TextInput(label="File Size", placeholder="e.g. 10GB", required=True, max_length=20)
    key       = discord.ui.TextInput(label="Decryption Key", placeholder="Enter decryption key", required=True, max_length=200)
    link      = discord.ui.TextInput(label="VIP Link", placeholder="https://mega.nz/...", required=True)
    image_url = discord.ui.TextInput(label="Image URL", placeholder="https://... (image URL)", required=True)

    async def on_submit(self, interaction: discord.Interaction):
        view = ChannelSelectView(
            post_name=self.post_name.value,
            file_size=self.file_size.value,
            key=self.key.value,
            link=self.link.value,
            image_url=self.image_url.value
        )
        await interaction.response.send_message("📢 Select a channel to post in:", view=view, ephemeral=True)


class ChannelSelectView(discord.ui.View):
    def __init__(self, post_name, file_size, key, link, image_url):
        super().__init__(timeout=180)
        self.add_item(ChannelSelect(post_name, file_size, key, link, image_url))


class ChannelSelect(discord.ui.Select):
    def __init__(self, post_name, file_size, key, link, image_url):
        self.post_name = post_name
        self.file_size = file_size
        self.key       = key
        self.link      = link
        self.image_url = image_url
        options = [discord.SelectOption(label=name, value=str(ch_id)) for name, ch_id in CHANNELS.items()]
        super().__init__(placeholder="Select channel...", options=options)

    async def callback(self, interaction: discord.Interaction):
        channel_id = int(self.values[0])
        channel    = client.get_channel(channel_id)
        if not channel:
            await interaction.response.send_message("❌ Channel not found.", ephemeral=True)
            return

        embed = discord.Embed(color=0x2b2d31)
        embed.set_image(url=self.image_url)
        embed.add_field(
            name=f"{self.post_name} — {self.file_size}",
            value=f"——————————————————\n🔗 **VIP Link:** ||{self.link}||\n\n**Decryption Key:** `{self.key}`\n——————————————————",
            inline=False
        )

        if isinstance(channel, discord.ForumChannel):
            thread, _ = await channel.create_thread(
                name=f"{self.post_name} — {self.file_size}",
                embed=embed,
            )
            await save_link(thread.id, self.link)
            await interaction.response.send_message("✅ Posted! Applying VIP window...", ephemeral=True)
            try:
                await reconcile_vip_window(channel)
            except Exception as e:
                print(f"[VIP reconcile] {e}")
        else:
            await channel.send(embed=embed)
            await interaction.response.send_message("✅ Posted successfully!", ephemeral=True)


class RevealLinkView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="Reveal Link", style=discord.ButtonStyle.primary, emoji="🔓", custom_id="reveal_link")
    async def reveal(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.defer(ephemeral=True)
        # 포럼 채널: channel_id = thread ID (저장 key와 일치)
        # 일반 채널: channel_id = channel ID ≠ msg.id → message.id로 폴백
        link = await get_link(interaction.channel_id)
        if not link:
            link = await get_link(interaction.message.id)
        if link:
            await interaction.followup.send(f"🔗 **Your VIP link:**\n{link}", ephemeral=True)
        else:
            await interaction.followup.send("❌ Link not found. Please contact an admin.", ephemeral=True)


class VIPRevealView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="Reveal Link (VIP Only)", style=discord.ButtonStyle.primary, emoji="🔒", custom_id="vip_reveal")
    async def reveal(self, interaction: discord.Interaction, button: discord.ui.Button):
        member = interaction.user
        has_vip = isinstance(member, discord.Member) and any(r.id == VIP_ROLE_ID_INT for r in member.roles)
        if not has_vip:
            await interaction.response.send_message(
                "🔒 This content is **VIP exclusive**.\nUpgrade to **VIP** at **xhouse.vip** to unlock the latest drops.",
                ephemeral=True
            )
            return
        link = await get_link(interaction.channel_id)
        if not link:
            link = await get_link(interaction.message.id)
        if link:
            key = link.split('#')[-1] if '#' in link else ''
            await interaction.response.send_message(
                f"🔗 **Your VIP link:**\n{link}\n\n**Decryption Key:** `{key}`",
                ephemeral=True
            )
        else:
            await interaction.response.send_message("❌ Link not found. Please contact an admin.", ephemeral=True)


# ── VIP 롤링 윈도우 헬퍼 ──
def _extract_link_key(value: str):
    """embed 필드 값에서 mega 링크와 키 추출"""
    link_m = re.search(r'\|\|(https://mega\.nz/[^\|]+)\|\|', value or '')
    key_m  = re.search(r'\*\*Decryption Key:\*\*\s*`([^`]+)`', value or '')
    link = link_m.group(1) if link_m else None
    key  = key_m.group(1) if key_m else (link.split('#')[-1] if link and '#' in link else '')
    return link, key


async def set_post_vip(thread: discord.Thread, vip: bool):
    """포스트를 VIP(버튼) 또는 공개(스포일러) 상태로 전환"""
    # 아카이브된 경우 임시 언아카이브
    was_archived = getattr(thread, 'archived', False)
    if was_archived:
        try:
            await thread.edit(archived=False)
            await asyncio.sleep(0.4)
        except Exception as e:
            return False, f"unarchive 실패: {e}"

    # 시작 메시지(임베드) + 별도 스포일러 메시지 수집
    starter = None
    try:
        starter = await thread.fetch_message(thread.id)
    except Exception:
        pass
    spoiler_msgs = []
    async for m in thread.history(limit=30, oldest_first=True):
        if m.author != client.user:
            continue
        if starter is None and m.embeds:
            starter = m
        if '||' in (m.content or '') and 'mega.nz' in (m.content or ''):
            spoiler_msgs.append(m)

    if not starter or not starter.embeds:
        return False, "임베드 메시지 없음"

    embed = starter.embeds[0]
    field = embed.fields[0] if embed.fields else None
    field_name = field.name if field else thread.name

    # 링크/키 확보: DB → 임베드 필드 → 별도 스포일러 메시지 순
    link = await get_link(thread.id)
    fkey = ''
    if field:
        flink, fkey = _extract_link_key(field.value)
        if not link:
            link = flink
    if not link:
        for sm in spoiler_msgs:
            slink, skey = _extract_link_key(sm.content)
            if slink:
                link = slink
                fkey = fkey or skey
                break
    if link:
        await save_link(thread.id, link)
    key = fkey or (link.split('#')[-1] if link and '#' in link else '')

    new_embed = discord.Embed(color=embed.color or 0x2b2d31)
    if embed.image and embed.image.url:
        new_embed.set_image(url=embed.image.url)

    if vip:
        new_value = (
            "——————————————————\n"
            "🔒 **VIP EXCLUSIVE**\n"
            "This drop is available to **VIP members** only.\n"
            "Tap the button below to unlock (VIP role required).\n"
            "Not VIP yet? Upgrade at **xhouse.vip**\n"
            "——————————————————"
        )
        new_embed.add_field(name=field_name, value=new_value, inline=False)
        await starter.edit(embed=new_embed, view=VIPRevealView())
        # 링크 노출되는 별도 스포일러 메시지 제거
        for sm in spoiler_msgs:
            if sm.id != starter.id:
                try:
                    await sm.delete()
                    await asyncio.sleep(0.3)
                except Exception:
                    pass
    else:
        if not link:
            return False, "복원할 링크 없음"
        new_value = (
            "——————————————————\n"
            f"🔗 **VIP Link:** ||{link}||\n\n"
            f"**Decryption Key:** `{key}`\n"
            "——————————————————"
        )
        new_embed.add_field(name=field_name, value=new_value, inline=False)
        await starter.edit(embed=new_embed, view=None)

    async with db_pool.acquire() as conn:
        await conn.execute(
            'INSERT INTO links (thread_id, mega_link, vip) VALUES ($1, $2, $3) '
            'ON CONFLICT (thread_id) DO UPDATE SET vip = $3',
            thread.id, link or '', vip
        )
    return True, "ok"


async def reconcile_vip_window(channel: discord.ForumChannel, limit: int = VIP_WINDOW_SIZE, deep: bool = False):
    """채널의 최신 limit개를 VIP로, 그 외 기존 VIP는 공개로 전환"""
    threads = {t.id: t for t in channel.threads}
    if deep:
        try:
            async for t in channel.archived_threads(limit=200):
                threads[t.id] = t
                await asyncio.sleep(0.4)
        except Exception:
            pass

    ordered = sorted(threads.values(), key=lambda t: t.id, reverse=True)
    vip_set = set(t.id for t in ordered[:limit])
    by_id   = {t.id: t for t in ordered}

    changes = []

    # 1) 최신 limit개 → VIP (아직 아니면)
    for t in ordered[:limit]:
        async with db_pool.acquire() as conn:
            r = await conn.fetchrow('SELECT vip FROM links WHERE thread_id=$1', t.id)
        if r and r['vip']:
            continue
        ok, _ = await set_post_vip(t, True)
        if ok:
            changes.append((t.id, True))
        await asyncio.sleep(1)

    # 2) 현재 VIP인데 윈도우 밖으로 밀린 것 → 공개
    async with db_pool.acquire() as conn:
        vip_rows = await conn.fetch('SELECT thread_id FROM links WHERE vip = TRUE')
    for row in vip_rows:
        tid = row['thread_id']
        if tid in vip_set:
            continue
        thread = by_id.get(tid)
        if thread is None:
            try:
                thread = await client.fetch_channel(tid)
            except Exception:
                # 스레드 삭제됨 → 플래그만 정리
                async with db_pool.acquire() as conn:
                    await conn.execute('UPDATE links SET vip = FALSE WHERE thread_id=$1', tid)
                continue
        if getattr(thread, 'parent_id', None) != channel.id:
            continue
        ok, _ = await set_post_vip(thread, False)
        if ok:
            changes.append((tid, False))
        await asyncio.sleep(1)

    return changes


class PostButtonView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="New Post", style=discord.ButtonStyle.success, emoji="📤", custom_id="new_post")
    async def new_post(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not interaction.user.guild_permissions.administrator:
            await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
            return
        await interaction.response.send_modal(PostModal())



# ================== Auto Post ==================
@tree.command(name="auto-post", description="엑셀 파일로 자동 포스팅 (A:채널 B:이름 C:Mega링크 D+:이미지URL)")
async def auto_post(interaction: discord.Interaction, file: discord.Attachment):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    # 엑셀 파일 읽기
    try:
        file_bytes = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        await interaction.followup.send(f"❌ Failed to read Excel file: {e}", ephemeral=True)
        return

    # 행 파싱 (A:채널 B:이름 C:Mega링크 D+:이미지URL)
    parsed = []
    parse_errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):  # 1행은 헤더
        if not any(row):
            continue

        channel_key = str(row[0]).strip().lower() if row[0] else ""
        folder_name = str(row[1]).strip() if row[1] else ""
        mega_link   = str(row[2]).strip() if row[2] else ""
        image_urls  = [str(c).strip() for c in row[3:] if c and str(c).strip()]

        if not channel_key or not folder_name or not mega_link:
            parse_errors.append(f"Row {i}: A/B/C 컬럼 필수 — 채널/이름/Mega링크")
            continue

        if channel_key not in CHANNEL_MAP:
            parse_errors.append(f"Row {i}: 알 수 없는 채널 `{channel_key}`")
            continue

        if not image_urls:
            parse_errors.append(f"Row {i}: `{folder_name}` — 이미지 URL 없음 (D열 이상 필요)")
            continue

        # Mega 링크에서 키 추출
        key = mega_link.split('#')[-1] if '#' in mega_link else ""

        parsed.append({
            "channel_key": channel_key,
            "folder_name": folder_name,
            "mega_link":   mega_link,
            "key":         key,
            "image_urls":  image_urls,
        })

    if not parsed:
        msg = "❌ No valid rows found.\n"
        if parse_errors:
            msg += "\n".join(parse_errors)
        await interaction.followup.send(msg, ephemeral=True)
        return

    # 중복 방지: DB에서 이미 포스팅된 이름 목록 조회
    async with db_pool.acquire() as conn:
        rows = await conn.fetch('SELECT name FROM posted_names')
    existing_names = {row['name'] for row in rows}

    success_list = []
    fail_list    = []
    skip_list    = []
    posted_channels = set()

    for item in parsed:
        folder_name = item["folder_name"]
        channel_key = item["channel_key"]
        mega_link   = item["mega_link"]
        key         = item["key"]
        image_urls  = item["image_urls"]
        channel_id  = CHANNEL_MAP[channel_key]
        channel     = client.get_channel(channel_id)

        if not channel:
            fail_list.append(f"❌ `{folder_name}` — Channel not found")
            continue

        # 중복 체크
        if folder_name in existing_names:
            skip_list.append(f"⏭️ `{folder_name}` — Already exists")
            continue

        try:
            embed = discord.Embed(color=0x2b2d31)
            embed.set_image(url=image_urls[0])
            embed.add_field(
                name=folder_name,
                value=f"——————————————————\n🔗 **VIP Link:** ||{mega_link}||\n\n**Decryption Key:** `{key}`\n——————————————————",
                inline=False
            )

            if isinstance(channel, discord.ForumChannel):
                thread, _ = await channel.create_thread(
                    name=folder_name,
                    embed=embed,
                )
                await save_link(thread.id, mega_link)
                posted_channels.add(channel.id)
                async with db_pool.acquire() as conn:
                    await conn.execute('INSERT INTO posted_names (name) VALUES ($1) ON CONFLICT DO NOTHING', folder_name)
                for extra_url in image_urls[1:]:
                    await thread.send(extra_url)
            else:
                await channel.send(embed=embed)
                async with db_pool.acquire() as conn:
                    await conn.execute('INSERT INTO posted_names (name) VALUES ($1) ON CONFLICT DO NOTHING', folder_name)
                for extra_url in image_urls[1:]:
                    await channel.send(extra_url)

            success_list.append(f"✅ `{folder_name}` → {channel_key}")

        except Exception as e:
            fail_list.append(f"❌ `{folder_name}` — {str(e)}")

    # VIP 윈도우 재조정 (포스팅된 채널만)
    for ch_id in posted_channels:
        ch = client.get_channel(ch_id)
        if isinstance(ch, discord.ForumChannel):
            try:
                await reconcile_vip_window(ch)
            except Exception as e:
                print(f"[VIP reconcile] {ch_id}: {e}")

    # 리포트
    report = f"📊 **Auto-Post Report**\n\n"

    if success_list:
        report += f"✅ **Success ({len(success_list)}):**\n"
        report += "\n".join(success_list) + "\n\n"

    if skip_list:
        report += f"⏭️ **Skipped — Already exists ({len(skip_list)}):**\n"
        report += "\n".join(skip_list) + "\n\n"

    if fail_list:
        report += f"❌ **Failed ({len(fail_list)}):**\n"
        report += "\n".join(fail_list) + "\n\n"

    if parse_errors:
        report += f"⚠️ **Parse Errors ({len(parse_errors)}):**\n"
        report += "\n".join(parse_errors)

    # Discord 2000자 제한 → 청크로 나눠서 전송
    chunks = []
    while len(report) > 1900:
        split_at = report.rfind('\n', 0, 1900)
        if split_at == -1:
            split_at = 1900
        chunks.append(report[:split_at])
        report = report[split_at:].lstrip('\n')
    chunks.append(report)

    for i, chunk in enumerate(chunks):
        await interaction.followup.send(chunk, ephemeral=True)

async def _run_migrate(notify_channel_id: int, notify_user_id: int):
    success = skipped = failed = 0

    for channel_id in CHANNEL_MAP.values():
        channel = client.get_channel(channel_id)
        if not channel or not isinstance(channel, discord.ForumChannel):
            continue

        all_threads = list(channel.threads)
        async for thread in channel.archived_threads(limit=None):
            all_threads.append(thread)
            await asyncio.sleep(1)  # 아카이브 페이지네이션 딜레이

        for thread in all_threads:
            try:
                link = await get_link(thread.id)
                if not link:
                    skipped += 1
                    await asyncio.sleep(3)
                    continue

                # 이미 링크 메시지가 전송된 스레드인지 확인
                already_done = False
                async for msg in thread.history(limit=20):
                    if msg.author == client.user and '||' in (msg.content or ''):
                        already_done = True
                        break
                await asyncio.sleep(1)

                if already_done:
                    skipped += 1
                    continue

                # 임베드 수정 대신 스레드에 링크 메시지 전송
                await thread.send(f"🔗 **VIP Link:** ||{link}||")
                success += 1
                print(f"[Migrate] ✅ {thread.name} ({success}번째)")
                await asyncio.sleep(10)  # edit 후 넉넉하게 대기

            except discord.HTTPException as e:
                if e.status == 429:
                    retry_after = float(getattr(e, 'retry_after', 10))
                    print(f"[Migrate] Rate limited — {retry_after}초 대기")
                    await asyncio.sleep(retry_after + 5)
                    failed += 1
                else:
                    print(f"[Migrate] thread {thread.id} 오류: {e}")
                    failed += 1
                    await asyncio.sleep(3)
            except Exception as e:
                print(f"[Migrate] thread {thread.id} 오류: {e}")
                failed += 1
                await asyncio.sleep(3)

    # 완료 후 실행자에게 DM 전송
    try:
        user = await client.fetch_user(notify_user_id)
        await user.send(f"✅ **Migration 완료**\n✅ 성공: {success}\n⏭️ 스킵: {skipped}\n❌ 실패: {failed}")
    except Exception:
        pass
    print(f"[Migrate] 완료 — 성공: {success}, 스킵: {skipped}, 실패: {failed}")


@tree.command(name="migrate-spoiler", description="기존 포스트 전부 스포일러 링크로 일괄 변환 (관리자 전용, 백그라운드 실행)")
async def migrate_spoiler(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return
    asyncio.create_task(_run_migrate(interaction.channel_id, interaction.user.id))
    await interaction.response.send_message(
        "⚙️ Migration 시작됨! 백그라운드에서 10초 간격으로 처리 중...\n완료되면 DM으로 결과 전송해줄게.",
        ephemeral=True
    )


@tree.command(name="update-links", description="엑셀로 Mega 링크 일괄 수정 (A:이름 B:새링크 C:스레드ID)")
async def update_links(interaction: discord.Interaction, file: discord.Attachment):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)

    try:
        file_bytes = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        await interaction.followup.send(f"❌ 파일 읽기 실패: {e}", ephemeral=True)
        return

    # A:이름(표시용) B:링크 C:스레드ID 파싱
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        link = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        raw_id = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        thread_id = raw_id.split('.')[0] if raw_id else ""  # float → int 처리
        if not link or not thread_id:
            continue
        if 'mega.nz' not in link:
            continue
        rows.append((name or thread_id, link, thread_id))

    if not rows:
        await interaction.followup.send("❌ 유효한 행이 없어. B열: 링크, C열: 스레드ID 확인해줘.", ephemeral=True)
        return

    notify_channel_id = interaction.channel_id
    notify_user_id = interaction.user.id
    asyncio.create_task(_run_update_links(rows, notify_channel_id, notify_user_id))
    await interaction.followup.send(
        f"⚙️ Update Links 시작! 백그라운드에서 {len(rows)}개 처리 중...\n완료되면 DM으로 결과 전송해줄게.",
        ephemeral=True
    )


async def _run_update_links(rows, notify_channel_id, notify_user_id):
    success_list = []
    fail_list = []

    for name, new_link, thread_id in rows:
        key = new_link.split('#')[-1] if '#' in new_link else ""

        try:
            matched = await client.fetch_channel(int(thread_id))
        except Exception as e:
            fail_list.append(f"❌ `{name}` — fetch 실패: {e}")
            continue

        try:
            # VIP 포스트는 링크가 공개 노출되면 안 됨 → DB만 갱신하고 VIP 상태 유지
            async with db_pool.acquire() as conn:
                vrow = await conn.fetchrow('SELECT vip FROM links WHERE thread_id=$1', matched.id)
            if vrow and vrow['vip']:
                await save_link(matched.id, new_link)
                await set_post_vip(matched, True)  # 임베드(이미지/제목)·버튼 갱신, 링크는 DB로만
                success_list.append(f"✅ `{name}` (VIP)")
                await asyncio.sleep(0.5)
                continue

            was_archived = getattr(matched, 'archived', False)
            if was_archived:
                await matched.edit(archived=False)
                await asyncio.sleep(0.5)

            embed_msg = None
            spoiler_msg = None
            async for msg in matched.history(limit=30, oldest_first=True):
                if msg.author != client.user:
                    continue
                if msg.embeds and embed_msg is None:
                    embed_msg = msg
                if '||' in (msg.content or '') and spoiler_msg is None:
                    spoiler_msg = msg

            new_content = f"🔗 **VIP Link:** ||{new_link}||"
            if key:
                new_content += f"\n**Decryption Key:** `{key}`"
            if spoiler_msg:
                await spoiler_msg.edit(content=new_content)
            else:
                await matched.send(new_content)
            await asyncio.sleep(0.5)

            if embed_msg and embed_msg.embeds:
                embed = embed_msg.embeds[0]
                new_embed = discord.Embed(color=embed.color or 0x2b2d31)
                if embed.image and embed.image.url:
                    new_embed.set_image(url=embed.image.url)
                for field in embed.fields:
                    new_value = field.value
                    new_value = re.sub(r'\|\|https://mega\.nz/[^\|]+\|\|', f'||{new_link}||', new_value)
                    new_value = re.sub(r'\*\*Decryption Key:\*\*\s*`[^`]*`', f'**Decryption Key:** `{key}`', new_value)
                    new_embed.add_field(name=field.name, value=new_value, inline=field.inline)
                if new_embed.fields:
                    await embed_msg.edit(embed=new_embed)
                    await asyncio.sleep(0.5)

            await save_link(matched.id, new_link)

            if was_archived:
                await matched.edit(archived=True)

            success_list.append(f"✅ `{name}`")
            await asyncio.sleep(0.5)

        except Exception as e:
            fail_list.append(f"❌ `{name}` — {e}")

    report = f"📊 **Update Links Report**\n\n"
    if success_list:
        report += f"✅ **성공 ({len(success_list)}):**\n" + "\n".join(success_list) + "\n\n"
    if fail_list:
        report += f"❌ **실패 ({len(fail_list)}):**\n" + "\n".join(fail_list)

    try:
        user = await client.fetch_user(notify_user_id)
        chunks = []
        while len(report) > 1900:
            split_at = report.rfind('\n', 0, 1900)
            if split_at == -1: split_at = 1900
            chunks.append(report[:split_at])
            report = report[split_at:].lstrip('\n')
        chunks.append(report)
        for chunk in chunks:
            await user.send(chunk)
    except Exception:
        pass


@tree.command(name="setup-post", description="관리자용 포스팅 버튼 설정")
async def setup_post(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
        return
    embed = discord.Embed(title="📤 Content Post Panel", description="Click the button below to post new content.", color=0x2b2d31)
    view = PostButtonView()
    await interaction.response.send_message(embed=embed, view=view)


# ================== 기존 포스트 링크 등록 ==================
@tree.command(name="set-link", description="기존 포스트에 링크 등록 (관리자 전용)")
async def set_link(interaction: discord.Interaction, thread_id: str, link: str):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
        return
    await save_link(int(thread_id), link)
    await interaction.response.send_message(f"✅ Link saved for thread `{thread_id}`!", ephemeral=True)


# ================== Support & Review System ==================
class StarSelect(discord.ui.Select):
    def __init__(self):
        options = [
            discord.SelectOption(label="⭐  1 Star",           value="1"),
            discord.SelectOption(label="⭐⭐  2 Stars",         value="2"),
            discord.SelectOption(label="⭐⭐⭐  3 Stars",       value="3"),
            discord.SelectOption(label="⭐⭐⭐⭐  4 Stars",     value="4"),
            discord.SelectOption(label="⭐⭐⭐⭐⭐  5 Stars",   value="5"),
        ]
        super().__init__(placeholder="Select your rating...", options=options)

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.send_modal(ReviewModal(stars=int(self.values[0])))


class StarRatingView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=120)
        self.add_item(StarSelect())


class ReviewModal(discord.ui.Modal, title="Leave a Review"):
    review_text = discord.ui.TextInput(
        label="Your Review",
        placeholder="Share your experience...",
        style=discord.TextStyle.paragraph,
        max_length=500
    )

    def __init__(self, stars: int):
        super().__init__()
        self.stars = stars

    async def on_submit(self, interaction: discord.Interaction):
        channel = client.get_channel(SUPPORT_CHANNEL_ID)
        if not channel:
            await interaction.response.send_message("❌ Error submitting review. Please contact an admin.", ephemeral=True)
            return

        stars_display = "⭐" * self.stars
        embed = discord.Embed(color=0x9b59b6, timestamp=datetime.utcnow())
        embed.set_author(name=interaction.user.display_name, icon_url=interaction.user.display_avatar.url)
        embed.description = f'*"{self.review_text.value}"*\n\n{stars_display}'
        embed.set_footer(text=f"{interaction.user.display_name} — Verified Member")

        await channel.send(f"<@{interaction.user.id}> left a review:", embed=embed)
        await interaction.response.send_message("✅ Thank you for your review! We really appreciate it.", ephemeral=True)


class SupportModal(discord.ui.Modal, title="Submit a Support Ticket"):
    subject = discord.ui.TextInput(
        label="Subject",
        placeholder="Brief description of your issue",
        max_length=100
    )
    message = discord.ui.TextInput(
        label="Message",
        placeholder="Describe your issue in detail...",
        style=discord.TextStyle.paragraph,
        max_length=1000
    )

    async def on_submit(self, interaction: discord.Interaction):
        channel = client.get_channel(SUPPORT_CHANNEL_ID)
        if not channel:
            await interaction.response.send_message("❌ Error submitting ticket. Please contact an admin.", ephemeral=True)
            return

        embed = discord.Embed(title="🎫 New Support Ticket", color=0x5865f2, timestamp=datetime.utcnow())
        embed.set_author(name=interaction.user.display_name, icon_url=interaction.user.display_avatar.url)
        embed.add_field(name="Subject", value=self.subject.value, inline=False)
        embed.add_field(name="Message", value=self.message.value, inline=False)
        embed.set_footer(text=f"{interaction.user} • {interaction.user.id}")

        await channel.send(f"<@{interaction.user.id}> submitted a support ticket:", embed=embed)
        await interaction.response.send_message("✅ Your ticket has been submitted! We'll get back to you as soon as possible.", ephemeral=True)


class SupportPanelView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)
        self.add_item(discord.ui.Button(label="Donate", style=discord.ButtonStyle.link, emoji="💝", url="https://buy.stripe.com/3cI6oJaZo8PS1kF9281ck0m"))

    @discord.ui.button(label="Support", style=discord.ButtonStyle.primary, emoji="🎫", custom_id="support_ticket")
    async def support(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(SupportModal())

    @discord.ui.button(label="Review", style=discord.ButtonStyle.secondary, emoji="⭐", custom_id="review_panel")
    async def review(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            "**Select your star rating:**",
            view=StarRatingView(),
            ephemeral=True
        )


@tree.command(name="setup-support", description="Support & Review 패널 설정")
async def setup_support(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ 관리자 권한이 필요합니다.", ephemeral=True)
        return

    embed = discord.Embed(
        title="💬 Support & Reviews",
        description="Need help or want to share your experience? Use the buttons below.",
        color=0x7b6eff
    )
    embed.add_field(name="🎫 Support", value="Having an issue? Submit a ticket and we'll help you out.", inline=False)
    embed.add_field(name="⭐ Review", value="Enjoying the content? Leave us a review — it means a lot!", inline=False)
    embed.add_field(name="💝 Donate", value="Want to support us? Any contribution is appreciated!", inline=False)
    view = SupportPanelView()
    await interaction.response.send_message(embed=embed, view=view)


# ================== 자동 Preview 포스팅 ==================
@tasks.loop(time=[
    time(0,  0, tzinfo=EDMONTON_TZ),   # 오전 12시 (자정)
    time(12, 0, tzinfo=EDMONTON_TZ),   # 오후 12시 (정오)
])
async def post_preview():
    channel = client.get_channel(PREVIEW_CHANNEL_ID)
    if not channel:
        return

    # 이전 메시지 삭제
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow('SELECT message_id FROM preview_msg ORDER BY id DESC LIMIT 1')
    if row:
        try:
            old_msg = await channel.fetch_message(row['message_id'])
            await old_msg.delete()
        except Exception:
            pass

    # 새 메시지 포스팅
    new_msg = await channel.send("@everyone\nCheck this free preview!  https://www.xhouse.vip/")

    # 메시지 ID 저장
    async with db_pool.acquire() as conn:
        await conn.execute('DELETE FROM preview_msg')
        await conn.execute('INSERT INTO preview_msg (message_id) VALUES ($1)', new_msg.id)


# ================== 프로모터 초대 설정 ==================
@tree.command(name="setup-promo-invite", description="프로모터 전용 초대 링크 생성")
async def setup_promo_invite(interaction: discord.Interaction, name: str):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    try:
        invite = await interaction.channel.create_invite(max_age=0, max_uses=0, unique=True, reason=f"Promoter invite: {name}")
        async with db_pool.acquire() as conn:
            await conn.execute(
                'INSERT INTO promoter_invites (promoter, invite_code, invite_url) VALUES ($1, $2, $3) ON CONFLICT (promoter) DO UPDATE SET invite_code=$2, invite_url=$3',
                name, invite.code, str(invite)
            )
        # 캐시 업데이트
        invites = await interaction.guild.invites()
        invite_cache[interaction.guild.id] = {inv.code: inv.uses for inv in invites}
        await interaction.followup.send(
            f"✅ Promoter invite created!\n**Name:** `{name}`\n**Link:** {invite}\n\n이제 `xhouse.vip/discount` 방문하면 이 링크로 자동 연결돼.",
            ephemeral=True
        )
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {e}", ephemeral=True)


# ================== 프로모터 통계 ==================
@tree.command(name="promo-stats", description="프로모터별 전환 통계")
async def promo_stats(interaction: discord.Interaction, promoter: str = None):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    async with db_pool.acquire() as conn:
        if promoter:
            rows = await conn.fetch(
                'SELECT ref, COUNT(*) AS cnt FROM referral_conversions WHERE ref=$1 GROUP BY ref',
                promoter
            )
        else:
            rows = await conn.fetch(
                'SELECT ref, COUNT(*) AS cnt FROM referral_conversions GROUP BY ref ORDER BY cnt DESC'
            )
    if not rows:
        await interaction.followup.send("No data yet.", ephemeral=True)
        return
    embed = discord.Embed(title="📊 Promoter Stats", color=0x7b6eff)
    for row in rows:
        embed.add_field(name=f"🔗 {row['ref'] or 'direct'}", value=f"**{row['cnt']}** conversions", inline=True)
    await interaction.followup.send(embed=embed, ephemeral=True)


# ================== 멤버 입장 (초대 추적 + 역할 자동 부여) ==================
@client.event
async def on_member_join(member: discord.Member):
    guild = member.guild
    if guild.id != XHOUSE_GUILD_ID_INT:
        return
    try:
        new_invites = await guild.invites()
    except Exception:
        return

    old_inv = invite_cache.get(guild.id, {})
    used_code = None
    for inv in new_invites:
        if inv.uses > old_inv.get(inv.code, 0):
            used_code = inv.code
            break

    invite_cache[guild.id] = {inv.code: inv.uses for inv in new_invites}

    if not used_code:
        return

    # ── 1. paid_invites 체크 (웹 결제 후 초대 플로우) ──
    async with db_pool.acquire() as conn:
        paid_row = await conn.fetchrow(
            'SELECT session_id, ref, used FROM paid_invites WHERE invite_code=$1', used_code
        )

    if paid_row and not paid_row['used']:
        role_granted = False
        role = guild.get_role(XHOUSE_ROLE_ID_INT)
        if role:
            try:
                await member.add_roles(role)
                role_granted = True
            except Exception as e:
                print(f"[Join] 역할 부여 실패: {e}")

        ref_label = paid_row['ref'] or 'direct'
        timestamp  = datetime.utcnow().strftime('%Y-%m-%d %H:%M')
        tx_channel = client.get_channel(TX_CHANNEL_ID)
        if tx_channel:
            status = "✅ Granted" if role_granted else "❌ FAILED"
            icon   = "✅" if role_granted else "⚠️"
            try:
                await tx_channel.send(f"{icon} Web Join | <@{member.id}> | ref: `{ref_label}` | Role: {status} | {timestamp} UTC")
            except Exception:
                pass

        if role_granted:
            try:
                await member.send("✅ Payment confirmed! Your membership role has been granted. Welcome to X-House! 🎉\n\n⭐ Enjoying your access? Drop a quick review in the server — it means a lot to us!")
            except Exception:
                pass

        async with db_pool.acquire() as conn:
            await conn.execute('UPDATE paid_invites SET used=TRUE WHERE invite_code=$1', used_code)
            await conn.execute(
                'UPDATE referral_conversions SET discord_id=$1 WHERE session_id=$2',
                member.id, paid_row['session_id']
            )
        return

    # ── 2. promoter_invites 체크 (Discord 결제 추적 플로우) ──
    async with db_pool.acquire() as conn:
        promo_row = await conn.fetchrow(
            'SELECT promoter FROM promoter_invites WHERE invite_code=$1', used_code
        )

    if promo_row:
        async with db_pool.acquire() as conn:
            await conn.execute(
                'INSERT INTO member_sources (discord_id, promoter) VALUES ($1, $2) ON CONFLICT (discord_id) DO NOTHING',
                member.id, promo_row['promoter']
            )
        print(f"[Join] Promoter source 기록: {member} → {promo_row['promoter']}")


# ================== VIP 윈도우 명령어 ==================
@tree.command(name="setup-vip-window", description="모든 채널 최신 20개를 VIP 전용으로 설정 (롤링 윈도우 초기화)")
async def setup_vip_window(interaction: discord.Interaction):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Administrator permission required.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    asyncio.create_task(_run_setup_vip_window(interaction.user.id))
    await interaction.followup.send(
        f"⚙️ VIP 윈도우 설정 시작! 각 채널 최신 {VIP_WINDOW_SIZE}개를 VIP 전용으로 전환 중...\n완료되면 DM으로 결과 보낼게.",
        ephemeral=True
    )


async def _run_setup_vip_window(notify_user_id: int):
    report = "📊 **VIP Window Setup**\n\n"
    for key, ch_id in CHANNEL_MAP.items():
        channel = client.get_channel(ch_id)
        if not isinstance(channel, discord.ForumChannel):
            report += f"⚠️ `{key}` — 포럼 채널 아님/없음\n"
            continue
        try:
            changes = await reconcile_vip_window(channel, deep=True)
            promoted = sum(1 for _, v in changes if v)
            demoted  = sum(1 for _, v in changes if not v)
            report += f"✅ `{key}` — VIP +{promoted} / 공개전환 {demoted}\n"
        except Exception as e:
            report += f"❌ `{key}` — {e}\n"
        await asyncio.sleep(1)

    try:
        user = await client.fetch_user(notify_user_id)
        await user.send(report)
    except Exception:
        pass


@tasks.loop(hours=6)
async def keep_vip_unarchived():
    """VIP 스레드가 아카이브되면 버튼이 안 먹으므로 주기적으로 깨움"""
    if db_pool is None:
        return
    async with db_pool.acquire() as conn:
        rows = await conn.fetch('SELECT thread_id FROM links WHERE vip = TRUE')
    vip_ids = {r['thread_id'] for r in rows}
    for ch_id in CHANNEL_MAP.values():
        channel = client.get_channel(ch_id)
        if not isinstance(channel, discord.ForumChannel):
            continue
        try:
            async for thread in channel.archived_threads(limit=50):
                if thread.id in vip_ids:
                    try:
                        await thread.edit(archived=False)
                        await asyncio.sleep(1)
                    except Exception:
                        pass
        except Exception:
            pass


# ================== 봇 시작 ==================
@client.event
async def on_ready():
    # ── 1. 영구 뷰 먼저 등록 (재시작 직후 버튼 클릭 즉시 응답 가능) ──
    client.add_view(RequestButtonView())
    client.add_view(PostButtonView())
    client.add_view(PaymentView())
    client.add_view(RevealLinkView())
    client.add_view(VIPRevealView())
    client.add_view(SupportPanelView())
    # ── 2. 이후 무거운 작업 처리 ──
    await init_db()
    await migrate_from_dummy()
    for guild in client.guilds:
        try:
            invites = await guild.invites()
            invite_cache[guild.id] = {inv.code: inv.uses for inv in invites}
        except Exception:
            pass
    await tree.sync()
    post_preview.start()
    if not keep_vip_unarchived.is_running():
        keep_vip_unarchived.start()
    print(f"✅ XHouse Bot 온라인! ({client.user})")

client.run(TOKEN)
